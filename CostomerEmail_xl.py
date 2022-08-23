from openpyxl import Workbook, load_workbook


class Customer_Email:
    def __init__(self):
        pass

    def Load_List(self):
        rows = []
        wb = load_workbook("Customer_Mail.xlsx")
        ws = wb.active
        for row in ws.iter_rows():
            rows.append(row[0].value)
        wb.close()
        return rows

    def RemoveEmail(self, email):
        rows = []
        count = 1
        wb = load_workbook("Customer_Mail.xlsx")
        ws = wb.active

        for row in ws.iter_rows():
            rows.append(row[0].value)

        for e in rows:
            if email != e:
                count += 1
            elif email == e:
                ws.delete_rows(count)

        wb.save("Customer_Mail.xlsx")
        wb.close()

    def SaveEmail(self, email):
        rows = []
        wb = load_workbook("Customer_Mail.xlsx")
        ws = wb.active
        if email == "":
            return
        for row in ws.iter_rows():
            rows.append(row[0].value)
        if len(rows) >= 1:
            for i in range(0, len(rows)):
                if rows[i] == email:
                    return
        ws.append([email])
        wb.save("Customer_Mail.xlsx")
        wb.close()

    def CreateEmail_xl(self):
        wb = Workbook()  # 새 워크북 생성
        ws = wb.active  # 현재 활성화된 sheet 가져옴

        ws.title = "emal"  # sheet 의 이름을 변경
        wb.save("Customer_Mail.xlsx")
        wb.close()
